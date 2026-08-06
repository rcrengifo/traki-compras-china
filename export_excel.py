"""
Genera un Excel de la cotizacion APROBADA para reenviar a China.
Incluye solo los productos aprobados, con la cantidad aprobada y sus fotos.
El texto va en ingles porque el destinatario es el proveedor en China.
"""
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

ROJO = "E30613"
NEGRO = "000000"


def _borde():
    fino = Side(style="thin", color="D9D9D9")
    return Border(left=fino, right=fino, top=fino, bottom=fino)


def generar(cot, lineas):
    """cot: dict de la cotizacion. lineas: lista de dicts (se filtran los Aprobados).
    Devuelve los bytes del .xlsx."""
    aprob = [l for l in lineas if l.get("estado") == "Aprobado"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved PO"

    # anchos de columna
    anchos = {"A": 6, "B": 46, "C": 16, "D": 12, "E": 8, "F": 12, "G": 14}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # --- titulo ---
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "ORDEN DE COMPRA APROBADA  /  APPROVED PURCHASE ORDER"
    t.font = Font(size=14, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill("solid", fgColor=NEGRO)
    ws.row_dimensions[1].height = 28

    # --- cabecera (bilingue Espanol / English) ---
    info = [
        ("Ref. / Quote No:", cot.get("referencia") or ""),
        ("Proveedor / Supplier:", cot.get("proveedor") or ""),
        ("Cliente / To (Buyer):", cot.get("cliente") or ""),
        ("Fecha / Date:", cot.get("fecha_emision") or ""),
        ("Incoterm:", cot.get("incoterm") or ""),
    ]
    r = 2
    for etiqueta, valor in info:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = etiqueta
        ws[f"A{r}"].font = Font(bold=True)
        ws.merge_cells(f"C{r}:G{r}")
        ws[f"C{r}"] = valor
        r += 1

    # --- encabezados de tabla (bilingue) ---
    hdr_row = r + 1
    headers = ["No.", "Producto / Product", "Imagen / Image", "Cant. aprob. / Approved Qty",
               "Unidad / Unit", "Precio unit / Unit Price", "Monto / Amount"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ROJO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde()

    # --- filas de productos ---
    total_amount = 0.0
    fila = hdr_row + 1
    for idx, l in enumerate(aprob, start=1):
        qty = l.get("cantidad_aprob")
        if qty is None:
            qty = l.get("cantidad")
        precio = l.get("precio_unit") or 0
        monto = (qty or 0) * precio
        total_amount += monto

        ws.cell(row=fila, column=1, value=idx)
        ws.cell(row=fila, column=2, value=l.get("descripcion"))
        # col 3 = imagen (se inserta despues)
        ws.cell(row=fila, column=4, value=qty)
        ws.cell(row=fila, column=5, value=l.get("unidad"))
        ws.cell(row=fila, column=6, value=precio)
        ws.cell(row=fila, column=7, value=round(monto, 2))
        for j in range(1, 8):
            cc = ws.cell(row=fila, column=j)
            cc.border = _borde()
            cc.alignment = Alignment(vertical="center", wrap_text=(j == 2))

        # insertar imagen si existe (bytes en la DB)
        img_bytes = l.get("imagen")
        if isinstance(img_bytes, (bytes, bytearray)) and len(img_bytes) > 0:
            try:
                bio = io.BytesIO(bytes(img_bytes))
                xi = XLImage(bio)
                xi.width, xi.height = 90, 90
                ws.row_dimensions[fila].height = 70
                ws.add_image(xi, f"C{fila}")
            except Exception:  # noqa: BLE001
                pass
        if ws.row_dimensions[fila].height is None:
            ws.row_dimensions[fila].height = 40
        fila += 1

    # --- total ---
    ws.cell(row=fila, column=6, value="TOTAL:").font = Font(bold=True)
    tc = ws.cell(row=fila, column=7, value=round(total_amount, 2))
    tc.font = Font(bold=True)
    tc.fill = PatternFill("solid", fgColor="F4F4F5")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

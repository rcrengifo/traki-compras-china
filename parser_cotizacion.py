"""
Lector de cotizaciones de proveedores de China (formato tipo "Quotation Sheet").

Extrae:
  - Cabecera: proveedor, contacto, cliente, fecha, incoterm, etc.
  - Lineas de producto: descripcion, cantidad, unidad, precio, total.
  - Imagenes embebidas, emparejadas a su fila/producto correcto.

Probado con el formato de Jiangsu Senmao Safety Technology, pero escrito
de forma tolerante para adaptarse a variaciones de otros proveedores.
"""
import os
import re
import zipfile
import unicodedata
import openpyxl


# ----- utilidades -------------------------------------------------------------

def _norm(texto):
    """Minusculas, sin acentos, sin espacios extra. Para comparar encabezados."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().lower()


# el código capturado debe contener al menos un dígito (evita falsos positivos como "Quotation Sheet")
_COD = r"((?=[\w\-/]*\d)[A-Za-z0-9][\w\-/]{2,})"
_REF_PATRONES = [
    r"\bP\.?\s*O\.?\s*(?:no|number)?\.?\s*[:：#]?\s*" + _COD,
    r"\b(?:quotation|quote|cotizaci[oó]n)\s*(?:no|number|n[°º]|#)?\.?\s*[:：#]?\s*" + _COD,
    r"\bref(?:erence|erencia)?\.?\s*(?:no)?\.?\s*[:：#]?\s*" + _COD,
    r"\border\s*(?:no)?\.?\s*[:：#]?\s*" + _COD,
    r"\bN[°º]\s*[:：#]?\s*" + _COD,
]


def _buscar_referencia(texto):
    """Busca un número de referencia / PO / cotización en el texto. None si no hay."""
    if not texto:
        return None
    for pat in _REF_PATRONES:
        m = re.search(pat, texto, re.I)
        if m:
            return m.group(1).strip(" .,-")
    return None


def _num(v):
    """Convierte a numero si se puede, si no devuelve None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = re.sub(r"[^\d.,\-]", "", str(v))
    if not s:
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


# ----- imagenes embebidas -----------------------------------------------------

def extraer_imagenes(ruta_xlsx, carpeta_destino, prefijo="img"):
    """
    Extrae las imagenes embebidas del xlsx y devuelve un dict {fila_excel: ruta_png}.
    La fila es 1-indexada (como se ve en Excel), tomada del ancla superior-izquierda.
    """
    os.makedirs(carpeta_destino, exist_ok=True)
    z = zipfile.ZipFile(ruta_xlsx)
    nombres = z.namelist()

    resultado = {}
    drawings = [n for n in nombres if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
    for draw_name in drawings:
        draw = z.read(draw_name).decode("utf-8", "ignore")
        rels_name = draw_name.replace("drawings/", "drawings/_rels/") + ".rels"
        if rels_name not in nombres:
            continue
        rels = z.read(rels_name).decode("utf-8", "ignore")
        rid2target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))

        anchors = re.findall(r"<xdr:twoCellAnchor.*?</xdr:twoCellAnchor>", draw, re.S)
        anchors += re.findall(r"<xdr:oneCellAnchor.*?</xdr:oneCellAnchor>", draw, re.S)
        for a in anchors:
            frm = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", a, re.S)
            rid = re.search(r'r:embed="(rId\d+)"', a)
            if not (frm and rid):
                continue
            fila_excel = int(frm.group(1)) + 1  # xdr:row es 0-index
            target = rid2target.get(rid.group(1), "")
            media = "xl/" + target.replace("../", "")
            if media not in nombres:
                continue
            ext = os.path.splitext(media)[1] or ".png"
            destino = os.path.join(carpeta_destino, f"{prefijo}_fila{fila_excel}{ext}")
            with open(destino, "wb") as f:
                f.write(z.read(media))
            resultado[fila_excel] = destino
    return resultado


# ----- lectura de la cotizacion ----------------------------------------------

# encabezados de columna que reconocemos (normalizados)
COLS = {
    "sn":          ["s/n", "sn", "no", "no.", "item", "#"],
    "descripcion": ["product description", "description", "descripcion", "producto"],
    "imagen":      ["image", "imagen", "picture", "foto"],
    "cantidad":    ["quantity", "qty", "cantidad"],
    "unidad":      ["unit", "unidad"],
    "precio_unit": ["unit price", "unit price(exw)", "precio unit", "precio unitario", "price"],
    "total":       ["total amount", "total", "amount", "monto"],
}


def _mapear_columnas(fila_encabezado):
    """Dada la fila de encabezados, devuelve {campo: indice_columna}."""
    mapa = {}
    for idx, celda in enumerate(fila_encabezado):
        n = _norm(celda)
        if not n:
            continue
        for campo, alias in COLS.items():
            if campo in mapa:
                continue
            if any(n == a or a in n for a in alias):
                mapa[campo] = idx
                break
    return mapa


def _extraer_cabecera(rows):
    """Busca proveedor, cliente, fecha, etc. recorriendo celda por celda.

    Trabaja por celda (y por linea dentro de la celda) con anclas ^ para no
    confundir, p.ej., el "to" dentro de la palabra "Town".
    """
    cab = {
        "referencia": None,
        "proveedor": None, "contacto": None, "email": None, "whatsapp": None,
        "cliente": None, "fecha_emision": None, "incoterm": None, "moneda": None,
        "lead_time": None, "forma_pago": None, "transporte": None, "empaque": None,
    }

    def set1(campo, valor):
        if valor and not cab.get(campo):
            cab[campo] = valor.strip(" .,-\t")

    # recorrer cada celda de las primeras 20 filas, y cada linea dentro de la celda
    for r in rows[:20]:
        for c in r:
            if c is None:
                continue
            # Proveedor/Cliente: buscar en la celda COMPLETA (a veces "From:" y el
            # nombre estan en lineas distintas de la misma celda).
            celda = re.sub(r"\s+", " ", str(c)).strip()
            m = re.match(r"from\s*[:：]\s*(.+?)(?:\s*add\s*[:：]|\s*whatsapp|\s*tel\s*[:：]|\s*e-?mail\s*[:：]|\s*contact\s*[:：]|$)", celda, re.I)
            if m and m.group(1).strip(" .,-"):
                set1("proveedor", m.group(1))
            m = re.match(r"to\s*[:：]\s*(.+)", celda, re.I)
            if m:
                set1("cliente", m.group(1))

            for linea in str(c).split("\n"):
                s = linea.strip()
                if not s:
                    continue

                m = re.search(r"issue date\s*[:：]?\s*([0-9]{1,2}[/\-][0-9]{1,2}[/\-][0-9]{2,4})", s, re.I)
                if m:
                    set1("fecha_emision", m.group(1))

                # Proveedor: "From: <nombre>" hasta Add/whatsapp/email/contact
                m = re.match(r"from\s*[:：]\s*(.+?)(?:\s*add\s*[:：]|\s*whatsapp|\s*tel[:：]|\s*e-?mail|\s*contact|$)", s, re.I)
                if m:
                    set1("proveedor", m.group(1))

                # Cliente: "To: <nombre>"  (anclado al inicio de la linea)
                m = re.match(r"to\s*[:：]\s*(.+)", s, re.I)
                if m:
                    set1("cliente", m.group(1))

                m = re.search(r"contact\s*[:：]\s*([^\n]+)", s, re.I)
                if m:
                    set1("contacto", m.group(1))
                m = re.search(r"([^\s:：]+@[^\s]+)", s)
                if m and "@" in m.group(1):
                    set1("email", m.group(1))
                m = re.search(r"(?:whatsapp|tel)\s*[/\s]*[:：]?\s*([+\d][\d\s\-]{6,})", s, re.I)
                if m:
                    set1("whatsapp", m.group(1))

                m = re.match(r"\d*\s*[.\-)]?\s*lead\s*time\s*[:：]\s*(.+)", s, re.I)
                if m:
                    set1("lead_time", m.group(1))
                m = re.match(r"\d*\s*[.\-)]?\s*payment\s*term\s*[:：]\s*(.+)", s, re.I)
                if m:
                    set1("forma_pago", m.group(1))
                m = re.match(r"\d*\s*[.\-)]?\s*transportation\s*[:：]\s*(.+)", s, re.I)
                if m:
                    set1("transporte", m.group(1))
                m = re.match(r"\d*\s*[.\-)]?\s*package\s*[:：]\s*(.+)", s, re.I)
                if m:
                    set1("empaque", m.group(1))

    # incoterm y moneda a partir de todo el texto
    texto = " ".join(str(c) for r in rows[:25] for c in r if c is not None).lower()
    for inc in ("exw", "fob", "cif", "cfr", "ddp", "dap"):
        if inc in texto:
            cab["incoterm"] = inc.upper()
            break
    if "$" in texto or "usd" in texto:
        cab["moneda"] = "USD"
    elif "rmb" in texto or "cny" in texto or "¥" in texto:
        cab["moneda"] = "CNY"

    # referencia / N° de cotizacion (texto sin bajar a minusculas, para preservar el codigo)
    texto_ref = " \n ".join(str(c) for r in rows[:25] for c in r if c is not None)
    cab["referencia"] = _buscar_referencia(texto_ref)
    return cab


def leer_cotizacion(ruta_xlsx, carpeta_imagenes=None):
    """
    Lee una cotizacion y devuelve dict con 'cabecera' y 'lineas'.
    Cada linea puede incluir 'imagen' (ruta al archivo extraido).
    """
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    cabecera = _extraer_cabecera(rows)

    # localizar la fila de encabezados de la tabla
    fila_hdr = None
    mapa = {}
    for i, r in enumerate(rows):
        m = _mapear_columnas(r)
        if "descripcion" in m and ("cantidad" in m or "precio_unit" in m):
            fila_hdr, mapa = i, m
            break
    if fila_hdr is None:
        return {"cabecera": cabecera, "lineas": [], "advertencia": "No se encontro la tabla de productos"}

    # imagenes por fila de excel
    imgs = {}
    if carpeta_imagenes:
        try:
            imgs = extraer_imagenes(ruta_xlsx, carpeta_imagenes)
        except Exception as e:  # noqa: BLE001
            imgs = {}
            cabecera["_img_error"] = str(e)

    lineas = []
    for i in range(fila_hdr + 1, len(rows)):
        r = rows[i]
        desc = r[mapa["descripcion"]] if mapa.get("descripcion") is not None and mapa["descripcion"] < len(r) else None
        desc = None if desc is None else str(desc).strip()
        # cortar en filas de totales / notas al pie
        if desc and re.match(r"^(exw total|total amount|\d+\.\s|lead time|payment|transportation|package)", _norm(desc)):
            break
        cantidad = _num(r[mapa["cantidad"]]) if mapa.get("cantidad") is not None and mapa["cantidad"] < len(r) else None
        if not desc and cantidad is None:
            continue
        fila_excel = i + 1  # iter_rows es 0-index -> excel 1-index
        lineas.append({
            "sn":          r[mapa["sn"]] if mapa.get("sn") is not None and mapa["sn"] < len(r) else len(lineas) + 1,
            "descripcion": desc or "",
            "cantidad":    cantidad,
            "unidad":      (str(r[mapa["unidad"]]).strip() if mapa.get("unidad") is not None and mapa["unidad"] < len(r) and r[mapa["unidad"]] else None),
            "precio_unit": _num(r[mapa["precio_unit"]]) if mapa.get("precio_unit") is not None and mapa["precio_unit"] < len(r) else None,
            "total":       _num(r[mapa["total"]]) if mapa.get("total") is not None and mapa["total"] < len(r) else None,
            "imagen":      imgs.get(fila_excel),
            "fila_excel":  fila_excel,
        })

    return {"cabecera": cabecera, "lineas": lineas}


if __name__ == "__main__":
    import sys, json
    ruta = sys.argv[1]
    data = leer_cotizacion(ruta, carpeta_imagenes="media_test")
    print("CABECERA:")
    for k, v in data["cabecera"].items():
        if v:
            print(f"  {k}: {v}")
    print(f"\nLINEAS ({len(data['lineas'])}):")
    for l in data["lineas"]:
        img = os.path.basename(l["imagen"]) if l["imagen"] else "(sin imagen)"
        print(f"  #{l['sn']}  {str(l['descripcion'])[:45]:45}  {l['cantidad']} {l['unidad'] or ''}  "
              f"x{l['precio_unit']} = {l['total']}  [{img}]")
